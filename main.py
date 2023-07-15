import os
import pyodbc
import pandas as pd
from datetime import datetime
import json
from openpyxl import load_workbook



caminho_pasta = "C:/Users/valer/Desktop/OneDrive/xlsx/"
caminho_processado = "Processado/"
caminho_rejeitado = "Rejeitado/"
caminho_logs = "logs/"
arquivo_log_basico = "log_basico.txt"
server = 'NOME SERVIDOR'
database = 'BANCO'
user = 'usuario'
password = 'senha'

conn_str = f"Driver={{SQL Server}};Server={server};Database={database};UID={user};PWD={password}"

# Não descomentar
# server = 'DESKTOP-SR3L8KN'
# database = 'teste'
# conn_str = f"Driver={{SQL Server}};Server={server};Database={database};Trusted_Connection=yes;"

# Função para verificar se o arquivo já existe e adicionar um número ao nome
def arquivo_existe(caminho, nome_arquivo):
    nome_base, extensao = os.path.splitext(nome_arquivo)
    novo_nome = nome_arquivo
    contador = 1

    while os.path.exists(os.path.join(caminho, novo_nome)):
        novo_nome = f"{nome_base}_{contador}{extensao}"
        contador += 1

    return novo_nome

conn = pyodbc.connect(conn_str)

arquivos_excel = [
    arquivo for arquivo in os.listdir(caminho_pasta) if arquivo.endswith(".xlsx")
]

log_existente = set()
if os.path.isfile(arquivo_log_basico):
    with open(arquivo_log_basico, "r") as log_file:
        log_existente = set(log_file.read().splitlines())

for arquivo_excel in arquivos_excel:
    if arquivo_excel in log_existente:
        print(f"O arquivo {arquivo_excel} já foi processado. Pulando para o próximo.")
        continue

    caminho_arquivo = os.path.join(caminho_pasta, arquivo_excel)
    df = pd.read_excel(caminho_arquivo, engine='openpyxl')
    col_sequencial_emissao = df["Sequencial"]
    col_cnpj = df["Cnpj"]
    linhas_atualizadas = []
    linhas_nao_atualizadas = []

    log_data = {
        "hora da conexão": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "arquivo importado": arquivo_excel,
        "linhas importadas": len(df),
        "linhas atualizadas": 0,
        "linhas rejeitadas": 0,
        "detalhes": {
            "colunas": df.columns.tolist(),
            "linhas": []
        }
    }

    for _, row in df.iterrows():
        sequencial_emissao = str(row["Sequencial"]).strip()
        cnpj = str(row["Cnpj"]).strip()
        tipo = str(row["Tipo"]).strip()

        # Processar de acordo com o tipo
        if tipo == "FRETE":
            contratante_query = f"SELECT id_Contratante FROM Contratante WHERE no_CNPJ in ('{cnpj}')"
            cursor = conn.cursor()
            cursor.execute(contratante_query)
            contratante_result = cursor.fetchall()

            if contratante_result:
                id_contratante = contratante_result[0][0]
                update_query = f"UPDATE Cartao SET id_Contratante = '{id_contratante}' WHERE no_SequencialEmissao = '{sequencial_emissao}' AND id_PessoaConta IS NULL AND id_tiposituacaocartao = '0'"
                cursor.execute(update_query)
                linhas_atualizadas.append(row)
                status = "Processada"
                log_data["linhas atualizadas"] += 1
            else:
                linhas_nao_atualizadas.append(row)
                status = "Rejeitada"
                log_data["linhas rejeitadas"] += 1

            print(f"Linha {status} - Sequencial: {sequencial_emissao}, CNPJ: {cnpj}")

            log_data["detalhes"]["linhas"].append({"Status": status, "Sequencial": sequencial_emissao, "CNPJ": cnpj})

        elif tipo == "GRUPO ECONOMICO":
            contratante_query = f"SELECT id_Contratante FROM Contratante WHERE no_CNPJ in ('{cnpj}')"
            cursor = conn.cursor()
            cursor.execute(contratante_query)
            contratante_result = cursor.fetchall()

            if contratante_result:
                update_query = f"UPDATE Cartao SET id_Contratante = NULL WHERE no_SequencialEmissao = '{sequencial_emissao}' AND id_PessoaConta IS NULL AND id_tiposituacaocartao = '0'"
                cursor.execute(update_query)
                linhas_atualizadas.append(row)
                status = "Processada"
                log_data["linhas atualizadas"] += 1
            else:
                linhas_nao_atualizadas.append(row)
                status = "Rejeitada"
                log_data["linhas rejeitadas"] += 1

            print(f"Linha {status} - Sequencial: {sequencial_emissao}, CNPJ: {cnpj}")

            log_data["detalhes"]["linhas"].append({"Status": status, "Sequencial": sequencial_emissao, "CNPJ": cnpj})

        elif tipo == "Corporativo":
            contratante_query = f"SELECT id_Contratante FROM Contratante WHERE no_CNPJ in ('{cnpj}')"
            cursor = conn.cursor()
            cursor.execute(contratante_query)
            contratante_result = cursor.fetchall()

            if contratante_result:
                id_contratante = contratante_result[0][0]
                produto_query = f"SELECT id_produtocartao FROM Produto_Cartao WHERE nm_ProdutoCartao LIKE '%{id_contratante}%'"
                cursor.execute(produto_query)
                produto_result = cursor.fetchall()

                if produto_result:
                    id_produtocartao = produto_result[0][0]
                    update_query = f"UPDATE Cartao SET id_Contratante = '{id_contratante}', id_produtocartao = '{id_produtocartao}' WHERE no_SequencialEmissao = '{sequencial_emissao}' AND id_PessoaConta IS NULL AND id_tiposituacaocartao = '0'"
                    cursor.execute(update_query)
                    linhas_atualizadas.append(row)
                    status = "Processada"
                    log_data["linhas atualizadas"] += 1
                else:
                    linhas_nao_atualizadas.append(row)
                    status = "Rejeitada"
                    log_data["linhas rejeitadas"] += 1
            else:
                linhas_nao_atualizadas.append(row)
                status = "Rejeitada"
                log_data["linhas rejeitadas"] += 1

            print(f"Linha {status} - Sequencial: {sequencial_emissao}, CNPJ: {cnpj}")

            log_data["detalhes"]["linhas"].append({"Status": status, "Sequencial": sequencial_emissao, "CNPJ": cnpj})

        else:
            # Tipo "Normal" - Fluxo padrão
            contratante_query = f"SELECT id_Contratante FROM Contratante WHERE no_CNPJ in ('{cnpj}')"
            cursor = conn.cursor()
            cursor.execute(contratante_query)
            contratante_result = cursor.fetchall()

            if contratante_result:
                id_contratante = contratante_result[0][0]
                update_query = f"UPDATE Cartao SET id_Contratante = '{id_contratante}' WHERE no_SequencialEmissao = '{sequencial_emissao}' AND id_PessoaConta IS NULL"
                cursor.execute(update_query)
                linhas_atualizadas.append(row)
                status = "Processada"
                log_data["linhas atualizadas"] += 1
            else:
                linhas_nao_atualizadas.append(row)
                status = "Rejeitada"
                log_data["linhas rejeitadas"] += 1

            print(f"Linha {status} - Sequencial: {sequencial_emissao}, CNPJ: {cnpj}")

            log_data["detalhes"]["linhas"].append({"Status": status, "Sequencial": sequencial_emissao, "CNPJ": cnpj})

    try:
        conn.commit()
        print(f"Arquivo {arquivo_excel} processado com sucesso!")

        arquivo_processado = arquivo_existe(caminho_processado, f"Processado_{arquivo_excel}")
        caminho_arquivo_processado = os.path.join(caminho_processado, arquivo_processado)
        df_atualizadas = pd.DataFrame(linhas_atualizadas)

        if os.path.isfile(caminho_arquivo_processado):
            with pd.ExcelWriter(caminho_arquivo_processado, mode="a", engine="openpyxl") as writer:
                writer.book = load_workbook(caminho_arquivo_processado)
                df_atualizadas.to_excel(writer, sheet_name="Sheet1", index=False, header=False)
        else:
            df_atualizadas.to_excel(caminho_arquivo_processado, index=False)

        arquivo_rejeitado = arquivo_existe(caminho_rejeitado, f"Rejeitado_{arquivo_excel}")
        caminho_arquivo_rejeitado = os.path.join(caminho_rejeitado, arquivo_rejeitado)
        df_nao_atualizadas = pd.DataFrame(linhas_nao_atualizadas)

        if os.path.isfile(caminho_arquivo_rejeitado):
            with pd.ExcelWriter(caminho_arquivo_rejeitado, mode="a", engine="openpyxl") as writer:
                writer.book = load_workbook(caminho_arquivo_rejeitado)
                df_nao_atualizadas.to_excel(writer, sheet_name="Sheet1", index=False, header=False)
        else:
            df_nao_atualizadas.to_excel(caminho_arquivo_rejeitado, index=False)

        with open(arquivo_log_basico, "a") as log_file:
            datan = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            log_file.write(f"Arquivo: {arquivo_excel} {datan}\n")
            log_file.write(f"Linhas processadas: {len(df)}\n")
            log_file.write(f"Linhas rejeitadas: {len(linhas_nao_atualizadas)}\n\n")

        nome_log = f"log_{arquivo_excel}.txt"
        caminho_log = os.path.join(caminho_logs, nome_log)

        if os.path.isfile(caminho_log):
            with open(caminho_log, "a") as log_file:
                log_file.write(json.dumps(log_data, indent=4))
        else:
            with open(caminho_log, "w") as log_file:
                log_file.write(json.dumps(log_data, indent=4))

        os.remove(caminho_arquivo)

    except Exception as e:
        print(f"Ocorreu um erro ao processar o arquivo {arquivo_excel}.")
        print(f"Erro: {str(e)}")

conn.close()
print("Conexão com o SQL fechada.")
